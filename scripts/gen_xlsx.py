#!/home/hermes/.venv/bin/python3
"""Generate an Excel file from /tmp/papers.json. Output: /tmp/ai-papers-summary.xlsx"""
import json, sys
from pathlib import Path

INPUT = "/tmp/papers.json"
OUTPUT = "/tmp/ai-papers-summary.xlsx"

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

papers = json.loads(Path(INPUT).read_text())
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Papers"
headers = ["Title", "Authors", "Year", "Summary"]
hfont = Font(bold=True)
hfill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hfont; cell.fill = hfill
for row, p in enumerate(papers, 2):
    ws.cell(row=row, column=1, value=p.get("title",""))
    ws.cell(row=row, column=2, value=p.get("authors",""))
    ws.cell(row=row, column=3, value=p.get("year",""))
    c = ws.cell(row=row, column=4, value=p.get("summary",""))
    c.alignment = Alignment(wrap_text=True)
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 80
ws.freeze_panes = "A2"
wb.save(OUTPUT)
print(f"XLSX written to {OUTPUT} ({len(papers)} papers)")
